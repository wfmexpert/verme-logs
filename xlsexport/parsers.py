import datetime
import re
from collections import Mapping, defaultdict
from io import BytesIO

import xlrd
from django.contrib.contenttypes.fields import GenericForeignKey
from django.contrib.postgres.fields.jsonb import JSONField
from django.db import transaction
from django.db.models import Q
from openpyxl import load_workbook


class CustomException(Exception):
    pass


class XLSParser:
    """
    Парсер xls-файлов
    """
    FORMAT = "xls"

    def __init__(self):
        self.cache_dict = dict()
        self.item_data = {}
        self.processed_items = set()
        self.created_items = defaultdict(list)

    def parse(self, template, file_contents):
        rb = xlrd.open_workbook(file_contents=file_contents)
        sheet = rb.sheet_by_index(0)
        errors = []
        for rownum in range(1, sheet.nrows):
            row = sheet.row_values(rownum)
            try:
                self.item_data = self.get_struct_from_row(row, rb, template, file_format=self.FORMAT)
                self.process_item_data(template)
            except Exception as exc:
                errors.append({'rownum': rownum, 'exc': exc, 'key': exc.key if hasattr(exc, 'key') else ''})
        return errors

    def get_attr_value(self, model, row_data, splitted_fields):
        # Определяем тип поля
        field = model._meta.get_field(splitted_fields[0])
        # Если поле - связь к другой модели
        if field.is_relation and field.related_model:
            # Обновляем модель
            model = field.related_model
            # Формируем массив для поиска объекта модели
            attr_query_dict = {splitted_fields[1]: self.item_data[row_data]}
        else:
            # Формируем массив для поиска объекта модели
            attr_query_dict = {splitted_fields[0]: self.item_data[row_data]}
        attr_value = model.objects.filter(**attr_query_dict).first()
        return attr_value

    def get_attr_value_ext(self, model, row_data, type_model=None):
        # Делим поле по разделителю
        splitted_fields = row_data.split('.')
        splitted_length = len(splitted_fields)

        current_idx = 0
        processed_model_list = list()
        processed_model_list.append(model)
        attr_value = None
        many_to_many_values = list()
        m2m_flag = False
        json_values = dict()
        fk_field = None

        while True:
            if not current_idx < splitted_length:
                break
            # Выбираем поле и модель
            current_field = splitted_fields[current_idx]
            try:
                field = model._meta.get_field(current_field)
            except Exception as error:
                error.key = current_field
                raise error
            # Если поле - связь к другой модели
            if field.is_relation:
                model = None
                if field.related_model:
                    # Обновляем модель
                    model = field.related_model
                elif type_model:
                    model = type_model
                    fk_field = field.fk_field
                    type_model = None
                if field.many_to_many:
                    m2m_flag = True
                if model:
                    processed_model_list.append(model)
            elif isinstance(field, JSONField):
                json_key = splitted_fields[current_idx+1]
                attr_value = self.item_data['.'.join(splitted_fields)]
                if current_field in json_values:
                    json_values[current_field][json_key] = attr_value
                else:
                    json_values[current_field] = {json_key: attr_value}
                break
            else:
                data_values = str(self.item_data[row_data]).split('|')
                if current_idx > 0:
                    if len(data_values) > 1:
                        for s_value in data_values:
                            # Формируем массив для поиска объекта модели
                            attr_query_dict = {splitted_fields[current_idx]: s_value}
                            # Получили объект модели по значению поля
                            attr_value = model.objects.filter(**attr_query_dict).first()
                            many_to_many_values.append(attr_value)
                        attr_value = many_to_many_values
                    else:
                        # Формируем массив для поиска объекта модели
                        attr_query_dict = {splitted_fields[current_idx]: self.item_data[row_data]}
                        # Получили объект модели по значению поля
                        try:
                            attr_value = model.objects.filter(**attr_query_dict).first()
                            if not attr_value:
                                raise Exception(f"Не обнаружен объект модели {model._meta.model_name} по значениям {list(attr_query_dict.values())}")
                        except Exception as error:
                            error.key = current_field
                            raise error
                        if m2m_flag:
                            many_to_many_values.append(attr_value)
                            attr_value = many_to_many_values
                else:
                    attr_value = self.item_data[row_data]
            current_idx += 1

        current_idx2 = splitted_length - 2
        # Убираем последнюю модель из списка обработанных
        # т.к. по ней уже имеем объект
        processed_model_list.pop()
        current_value = attr_value

        while True:
            if current_idx2 < 1:
                break
            current_model = processed_model_list.pop()
            #
            current_field = splitted_fields[current_idx2]
            field = current_model._meta.get_field(current_field)
            #
            if not field.many_to_many:
                current_object_attr_query = {splitted_fields[current_idx2]: current_value}

                if processed_model_list:
                    current_value = current_model.objects.filter(**current_object_attr_query).first()
                else:
                    current_value = current_object_attr_query
            # ManytoMany Field
            else:
                search_values = current_value.split('|')
                for s_value in search_values:
                    current_object_attr_query = {splitted_fields[current_idx2]: s_value}
                    if processed_model_list:
                        current_value = current_model.objects.filter(**current_object_attr_query).first()
                    else:
                        current_value = current_object_attr_query
                    many_to_many_values.append(current_value)

            current_idx2 -= 1

        result_dict = dict()
        m2m_dict = dict()
        field_name = fk_field or splitted_fields[0]
        if many_to_many_values:
            m2m_dict = {field_name: many_to_many_values}
        elif field_name not in json_values:
            if splitted_length > 1:
                value = current_value
            else:
                value = attr_value
            if fk_field:
                value = value.id
            result_dict = {field_name: value}

        return result_dict, m2m_dict, json_values

    def process_item_data(self, template):
        model = template.get_model()
        param_fields, fields = template.get_param_fields()
        key_fields = template.get_key_fields()

        if not param_fields or not key_fields:
            raise CustomException(f"Не указаны поля или ключевые поля в шаблоне")

        # Список ключевых полей
        key_fields_list = set()
        for key_field in key_fields:
            key_fields_list.add(key_field['field'])

        # Список игнорируемых полей
        ignore_fields_list = set()
        for param_field in param_fields:
            if param_field.get('import_ignore', True):
                ignore_fields_list.add(param_field['field'])
        # Список параметризированных полей и проверка GenericFields
        field_names = {field.name: field for field in fields}
        base_field_names = [param_field["field"].split(".")[0] for param_field in param_fields]
        generic_fk_fields = {field.fk_field: field for field in fields if isinstance(field, GenericForeignKey)}
        parameterized_fields = dict()
        for param_field in param_fields:
            if param_field["field"] in ignore_fields_list:
                continue
            base_name = param_field["field"].split(".")[0]
            if base_name in field_names and isinstance(field_names[base_name], GenericForeignKey):
                parameterized_fields[param_field["field"]] = field_names[base_name].ct_field
                if parameterized_fields[param_field["field"]] not in base_field_names:
                    raise Exception(f"Не обнаружен тип поля {param_field['field']} в параметрах")
            if base_name in generic_fk_fields and generic_fk_fields[base_name].ct_field not in base_field_names:
                raise Exception(f"Не обнаружен тип поля {param_field['field']} в параметрах")

        # Для update_or_create
        result_query = dict()

        def dict_merge(dct, merge_dct):
            for k, v in merge_dct.items():
                if (k in dct and isinstance(dct[k], dict)
                        and isinstance(merge_dct[k], Mapping)):
                    dict_merge(dct[k], merge_dct[k])
                else:
                    dct[k] = merge_dct[k]

        query = dict()
        defaults = dict()
        m2m = dict()
        cache_set = list()
        json_fields_to_update = dict()
        for row_data in self.item_data:
            if row_data in ignore_fields_list:
                continue
            type_model = None
            if row_data in parameterized_fields:
                type_field = defaults.get(parameterized_fields[row_data])
                if not type_field:
                    type_field = query.get(parameterized_fields[row_data])
                if type_field:
                    type_model = type_field.model_class()
            # Если поле ключевое
            if row_data in key_fields_list:
                cache_set.append(self.item_data[row_data])
                query_dict, m2m_dict, json_dict = self.get_attr_value_ext(model, row_data, type_model)
                query.update(query_dict)
                m2m.update(m2m_dict)
            else:
                # Если поле не ключевое
                query_dict, m2m_dict, json_dict = self.get_attr_value_ext(model, row_data, type_model)
                defaults.update(query_dict)
                m2m.update(m2m_dict)
                dict_merge(json_fields_to_update, json_dict)

        target_object = None
        if cache_set:
            cache_tuple = tuple(i for i in cache_set)
            if cache_tuple in self.cache_dict:
                target_object = self.cache_dict.get(cache_tuple)
            else:
                target_object = model.objects.filter(**query).first()
                self.cache_dict.update({cache_tuple: target_object})

        with transaction.atomic():
            def compile_expression(expression):
                return compile(expression, f'm2m_set', 'exec')

            def exec_expression(expression, target_object, m2m=None):
                locals_ = {
                    'target_object': target_object,
                    'm2m': m2m
                }
                exec(compile_expression(expression), None, locals_)
                return locals_.get('result', None)
            try:
                error_key = ''
                if target_object:
                    for key, value in defaults.items():
                        error_key = key
                        setattr(target_object, key, value)
                    # Установка значений JSON полей
                    for k, v in json_fields_to_update.items():
                        json_field = getattr(target_object, k)
                        if not json_field:
                            json_field = dict()
                            setattr(target_object, k, json_field)
                        dict_merge(json_field, v)
                        error_key = k
                    target_object.save()

                    # Установка M2M полей
                    for m2m_key, m2m_value in m2m.items():
                        error_key = m2m_key
                        expression = f'target_object.{m2m_key}.set(m2m)'
                        exec_expression(expression, target_object, m2m_value)
                else:
                    result_query.update(query)
                    result_query.update(defaults)
                    result_query.update(json_fields_to_update)

                    target_object = model.objects.create(**result_query)
                    if target_object.pk is None:
                        # Объект создан в "партицированной" таблице и его нужно выбрать заново т.к.
                        # триггер, разбарсывающий записи по таблицам возвращает NULL
                        target_object = model.objects.filter(**result_query).order_by("-id").first()                    

                    # Установка M2M полей
                    for m2m_key, m2m_value in m2m.items():
                        error_key = m2m_key
                        expression = f'target_object.{m2m_key}.set(m2m)'
                        exec_expression(expression, target_object, m2m_value)
                    self.created_items[target_object] = self.item_data
            except Exception as error:
                error.key = error_key
                raise error
        self.processed_items.add(target_object)

    @staticmethod
    def get_struct_from_row(row, rb, template, file_format):

        def get_formatted_field(value, format, file_format):
            if format.find(".0") != -1:
                return get_float(value, len(format.split(".")[-1].strip()))
            elif format.find("#") != -1:
                return get_int(value)
            elif format == "@":
                return str(value)
            elif format.startswith("%"):
                try:
                    formatted_value = datetime.datetime.strptime(value, format)
                except TypeError:
                    formatted_value = get_cell_date(value, file_format)
                return formatted_value
            else:
                return get_cell_date(value, file_format)

        def get_cell_date(cell, file_format):
            try:
                if file_format == "xls":
                    return str(cell).strip() and datetime.datetime(*xlrd.xldate_as_tuple(cell, rb.datemode)) or None
                else:
                    return str(cell).strip() or None
            except TypeError:
                pass

        def get_float(value, prec=2):
            try:
                return round(float(value), prec)
            except (TypeError, ValueError):
                pass

        def get_int(value):
            try:
                return int(value)
            except (TypeError, ValueError):
                pass

        def clean_value(value):
            if isinstance(value, (bytes, str)):
                if value:
                    if isinstance(value, bytes):
                        value = value.encode('utf-8')
                    value = value.strip().replace("'", '')
                else:
                    value = ''
            return value

        result = dict()
        param_fields, fields = template.get_param_fields()
        for idx, item in enumerate(row):
            param_field = param_fields[idx]
            field_path = param_field['field']
            model_field = None
            if field_path.count('.') == 0:
                model_field = template.get_model()._meta.get_field(field_path)
            attr_value = clean_value(row[idx])
            value = None
            if 'format' in param_field and attr_value:
                value = get_formatted_field(attr_value, param_field['format'], file_format)
                if not value and attr_value:
                    value = attr_value
            elif model_field and hasattr(model_field, "get_internal_type") and model_field.get_internal_type() == 'DurationField':
                # если это простое поле типа min_value (а не table.code), и в модели там хранится продолжительность, делаем продолжительность
                value = datetime.timedelta(minutes=int(row[idx]) if row[idx] else 0)
            elif attr_value:
                value = attr_value
            result.update({field_path: value})
        return result


class XLSXParser(XLSParser):
    FORMAT = "xlsx"

    def parse(self, template, file_contents):
        rb = load_workbook(BytesIO(file_contents))
        sheet = rb.worksheets[0]
        errors = []
        for rownum, row in enumerate(sheet.iter_rows(values_only=True)):
            if not rownum:
                continue
            try:
                self.item_data = self.get_struct_from_row(row, rb, template, file_format=self.FORMAT)
                self.process_item_data(template)
            except Exception as exc:
                errors.append({'rownum': rownum, 'exc': exc, 'key': exc.key if hasattr(exc, 'key') else ''})
        return errors
